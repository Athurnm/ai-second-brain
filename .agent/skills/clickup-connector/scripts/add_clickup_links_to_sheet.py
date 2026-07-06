#!/usr/bin/env python3
"""
Adds a 'ClickUp Link' column to the FINA PoC Action Item Google Sheet.
Strategy: export as XLSX → add column → re-upload converting back to Google Sheets.
"""

import io
import json
import sys
import time
import requests
import openpyxl

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

CLICKUP_TOKEN = "pk_88954348_1FB6GNJATX3QOVXKP38TABFTA3U9Q0XS"
SHEET_FILE_ID = "<YOUR_DRIVE_ID>"
TOKEN_PATH    = ".agent/skills/secondary-drive-connector/token.json"

CATEGORY_PREFIX = {
    "Pre-Build / Alignment":       "[Pre-Build]",
    "Infrastructure":              "[Infra]",
    "Backend - Orchestrator Core": "[BE-Core]",
    "Backend - Scenario Flows":    "[BE-Flow]",
    "Backend - API Layer":         "[BE-API]",
    "Tool Registry":               "[Tools]",
    "RAG / Knowledge Base":        "[RAG]",
    "Configuration":               "[Config]",
    "Analytics / Measurement":     "[Analytics]",
    "Testing / QA":                "[QA]",
    "PM / Product":                "[PM]",
}

def get_drive_service():
    with open(TOKEN_PATH) as f:
        td = json.load(f)
    creds = Credentials(
        token=td["token"],
        refresh_token=td["refresh_token"],
        token_uri=td["token_uri"],
        client_id=td["client_id"],
        client_secret=td["client_secret"],
        scopes=td["scopes"],
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("drive", "v3", credentials=creds), creds

def fetch_clickup_tasks():
    """Return dict: clickup_task_name_without_prefix -> clickup_url"""
    headers = {"Authorization": CLICKUP_TOKEN}
    url = f"https://api.clickup.com/api/v2/list/901614365044/task"
    tasks = {}
    page = 0
    while True:
        r = requests.get(url, headers=headers,
                         params={"page": page, "include_closed": "true"}, timeout=30)
        r.raise_for_status()
        data = r.json()
        for t in data.get("tasks", []):
            name = t["name"]
            # Strip prefix like "[Pre-Build] " to get bare action item
            bare = name
            for prefix in CATEGORY_PREFIX.values():
                if name.startswith(prefix + " "):
                    bare = name[len(prefix) + 1:]
                    break
            tasks[bare] = t.get("url", f"https://app.clickup.com/t/{t['id']}")
        if data.get("last_page", True):
            break
        page += 1
    print(f"Fetched {len(tasks)} tasks from ClickUp.")
    return tasks

def export_as_xlsx(service):
    print("Exporting Google Sheet as XLSX...")
    content = service.files().export(
        fileId=SHEET_FILE_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ).execute()
    print(f"  Exported {len(content):,} bytes.")
    return io.BytesIO(content)

def add_clickup_column(xlsx_bytes, clickup_tasks):
    """
    Opens the XLSX, finds the target sheet (gid 1924263051 → first matching sheet),
    adds/updates a 'ClickUp Link' column (column I, index 9).
    Returns modified workbook as bytes.
    """
    wb = openpyxl.load_workbook(xlsx_bytes)

    # The target sheet — try by name first, fallback to first sheet
    target_sheet = None
    for ws in wb.worksheets:
        if "Action Item" in ws.title or "FINA" in ws.title or "PoC" in ws.title:
            target_sheet = ws
            break
    if target_sheet is None:
        target_sheet = wb.worksheets[0]

    print(f"  Working on sheet: '{target_sheet.title}'")
    print(f"  Rows: {target_sheet.max_row}, Cols: {target_sheet.max_column}")

    # Find or create 'ClickUp Link' column
    header_row = list(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    print(f"  Headers: {header_row}")

    link_col = None
    for idx, cell in enumerate(header_row, start=1):
        if cell and "clickup" in str(cell).lower():
            link_col = idx
            break

    if link_col is None:
        link_col = target_sheet.max_column + 1
        target_sheet.cell(row=1, column=link_col, value="ClickUp Link")
        print(f"  Added 'ClickUp Link' header at column {link_col}.")
    else:
        print(f"  'ClickUp Link' column already exists at column {link_col}.")

    # Find Action Item column (column C = index 3)
    action_col = None
    for idx, cell in enumerate(header_row, start=1):
        if cell and str(cell).strip().lower() in ("action item", "action_item"):
            action_col = idx
            break
    if action_col is None:
        action_col = 3  # fallback to column C
    print(f"  Action Item column: {action_col}")

    # Populate ClickUp links
    matched, unmatched = 0, []
    for row in target_sheet.iter_rows(min_row=2, max_row=target_sheet.max_row):
        action_cell = row[action_col - 1]
        action_item = action_cell.value
        if not action_item:
            continue

        cu_url = clickup_tasks.get(str(action_item).strip())
        link_cell = target_sheet.cell(row=action_cell.row, column=link_col)
        if cu_url:
            link_cell.value = cu_url
            matched += 1
        else:
            link_cell.value = "— not found —"
            unmatched.append(str(action_item)[:60])

    print(f"  Matched: {matched} | Unmatched: {len(unmatched)}")
    if unmatched:
        print("  Unmatched rows:")
        for u in unmatched:
            print(f"    - {u}")

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def reupload_as_gsheet(service, xlsx_bytes):
    """Re-upload the XLSX as a Google Sheet (replace the existing file)."""
    print("Re-uploading as Google Sheet...")
    media = MediaIoBaseUpload(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    updated = service.files().update(
        fileId=SHEET_FILE_ID,
        media_body=media,
        # Convert back to Google Sheets format
        body={},
        supportsAllDrives=True,
    ).execute()
    print(f"  Upload complete. File ID: {updated.get('id')}")
    return updated

def main():
    service, creds = get_drive_service()

    # 1. Fetch ClickUp tasks
    clickup_tasks = fetch_clickup_tasks()

    # 2. Export sheet as XLSX
    xlsx_io = export_as_xlsx(service)

    # 3. Add ClickUp Link column
    modified_xlsx = add_clickup_column(xlsx_io, clickup_tasks)

    # 4. Re-upload back to Google Drive as Google Sheets
    reupload_as_gsheet(service, modified_xlsx)

    print("\nDone! Check the sheet for the new 'ClickUp Link' column.")
    print(f"https://docs.google.com/spreadsheets/d/{SHEET_FILE_ID}/edit")

if __name__ == "__main__":
    main()
